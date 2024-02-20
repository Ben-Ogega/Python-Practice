import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# print(product_list.max_row) 75 rows including the header

''' 
    Calculation for total value of inventory per supplier 
'''
for product_row in range(2, product_list.max_row + 1):

    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in product_per_supplier:

        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1

    else:
        product_per_supplier[supplier_name] = 1

    ''' 
        Calculation for total value of inventory per supplier 
    '''

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price

    else:
        total_value_per_supplier[supplier_name] = inventory * price

    ''' 
           Calculate products with inventory less than 10 
       '''

    if inventory < 10:
        products_under_10_inv[product_number] = inventory

    # Add value for total inventory price
    inventory_price.value = inventory * price

    # Convert dictionary keys to a list
keys_values = list(product_per_supplier.keys())

# Check if the index is within the range of the keys
for key in keys_values:
    # print(key)
    if key in product_per_supplier:
        value = product_per_supplier[key]
        print(f'{key}: supplied us with {value} products')

    else:
        print(f'{key}: Key not found in product_per_supplier')

print('\n',total_value_per_supplier)
print(f'The following products {products_under_10_inv} have an inventory of less than 10')
inv_file.save('updated_inventory.xlsx')